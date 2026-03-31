import os, random, zipfile
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference

random.seed(2026) # random seed for reproducibility
np.random.seed(2026)

base = Path("latvia_gov_excel_datasets")
if base.exists():
    import shutil
    shutil.rmtree(base)

folders = [
    base / "1_budget_reports",
    base / "2_procurement_data",
    base / "3_employee_registry",
    base / "4_municipality_data",
    base / "5_mixed_quality_data",
    base / "6_reporting_template",
    base / "7_chart_data",
]
for f in folders:
    f.mkdir(parents=True, exist_ok=True)

municipalities = [
    "Rīga", "Daugavpils", "Liepāja", "Jelgava", "Jūrmala", "Ventspils",
    "Rēzekne", "Valmiera", "Ogre", "Cēsis", "Sigulda", "Tukums", "Kuldīga",
    "Talsi", "Bauska", "Saldus", "Dobele", "Madona", "Ludza", "Alūksne"
]

departments_lv = [
    "IT nodaļa", "Personāla nodaļa", "Finanšu nodaļa", "Iepirkumu nodaļa",
    "Juridiskā nodaļa", "Klientu apkalpošana", "Attīstības nodaļa", "Izglītības pārvalde"
]

categories_lv = {
    "IT nodaļa": ["Programmatūra", "Datorsistēmas", "Licences", "Apkope"],
    "Personāla nodaļa": ["Apmācības", "Personāla atlase", "Komandējumi"],
    "Finanšu nodaļa": ["Konsultācijas", "Revīzija", "Biroja izdevumi"],
    "Iepirkumu nodaļa": ["Kancelejas preces", "Pakalpojumi", "Līgumu administrēšana"],
    "Juridiskā nodaļa": ["Juridiskie pakalpojumi", "Dokumentu sagatavošana"],
    "Klientu apkalpošana": ["Drukāšana", "Zvanu centrs", "Materiāli"],
    "Attīstības nodaļa": ["Projekti", "Semināri", "Pētījumi"],
    "Izglītības pārvalde": ["Skolu atbalsts", "Mācību materiāli", "Pasākumi"],
}

names_first = ["Jānis", "Anna", "Mārtiņš", "Ilze", "Andris", "Līga", "Edgars", "Zane", "Agnese", "Kārlis", "Laura", "Māris", "Kristīne", "Toms", "Elīna"]
names_last = ["Bērziņš", "Kalniņa", "Ozols", "Liepiņa", "Krūmiņš", "Pētersone", "Eglītis", "Siliņa", "Vītols", "Aboliņa", "Rudzītis", "Mežs", "Jansons", "Dreimane"]

def synth_name():
    return f"{random.choice(names_first)} {random.choice(names_last)}"

def save_df_xlsx(df, path, sheet_name="Dati"):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 28)

# 1. Budget reports
months = [
    ("2025-01", "janvāris"),
    ("2025-02", "februāris"),
    ("2025-03", "marts"),
    ("2025-04", "aprīlis"),
    ("2025-05", "maijs"),
    ("2025-06", "jūnijs"),
]
budget_summary_rows = []

for ym, month_lv in months:
    year, month = map(int, ym.split("-"))
    rows = []
    for _ in range(42):
        dep = random.choice(departments_lv)
        cat = random.choice(categories_lv[dep])
        amt = round(np.random.uniform(75, 6500), 2)
        day = random.randint(1, 28)
        dt = pd.Timestamp(year=year, month=month, day=day)
        date_value = random.choice([
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%d.%m.%Y"),
            dt.strftime("%d/%m/%Y"),
            dt
        ])
        if random.random() < 0.06:
            amt = None
        if random.random() < 0.05:
            cat = None

        municipality = random.choice(municipalities[:10])
        rows.append({
            "Iestāde": municipality + " pašvaldība",
            "Department": dep,
            "Category": cat,
            "Amount EUR": amt,
            "Date": date_value,
            "Invoice No": f"INV-{year}{month:02d}-{random.randint(1000,9999)}"
        })

        if amt is not None:
            budget_summary_rows.append({
                "Month": month_lv.capitalize(),
                "Department": dep,
                "Amount": amt
            })

    save_df_xlsx(pd.DataFrame(rows), base / "1_budget_reports" / f"budget_{ym}.xlsx", sheet_name="Izdevumi")

# 2. Procurement data
supplier_canonical = [
    "SIA BaltTech risinājumi",
    "SIA Datu Serviss",
    "SIA Ziemeļu Birojs",
    "SIA E-pārvalde",
    "SIA Nova Sistēmas",
    "SIA Viedā Vide",
    "SIA Latgales Projekti",
]
supplier_variants = {
    "SIA BaltTech risinājumi": ['SIA "BaltTech risinājumi"', "BaltTech risinajumi SIA", "SIA BaltTech", "Baltech Risinajumi"],
    "SIA Datu Serviss": ['SIA "Datu Serviss"', "Datu Serviss", "DATU SERVISS SIA"],
    "SIA Ziemeļu Birojs": ['SIA "Ziemeļu Birojs"', "Ziemelu Birojs SIA", "SIA Ziemeļu birojs"],
    "SIA E-pārvalde": ['SIA "E-pārvalde"', "E-parvalde", "SIA E parvalde"],
    "SIA Nova Sistēmas": ['SIA "Nova Sistēmas"', "Nova Sistemas", "SIA NOVA SISTĒMAS"],
    "SIA Viedā Vide": ['SIA "Viedā Vide"', "Vieda Vide", "SIA VIEDA VIDE"],
    "SIA Latgales Projekti": ['SIA "Latgales Projekti"', "Latgales Projekti", "LATGALES PROJEKTI SIA"],
}
contract_types = ["IT", "Būvniecība", "Apmācības", "Juridiskie pakalpojumi", "Kancelejas preces", "Pētījumi"]
authorities = [m + " valstspilsētas pašvaldība" if m in ["Rīga","Daugavpils","Liepāja","Jelgava","Jūrmala","Ventspils","Rēzekne"] else m + " novada pašvaldība" for m in municipalities[:12]]

rows = []
for _ in range(120):
    canon = random.choice(supplier_canonical)
    supplier = random.choice(supplier_variants[canon])
    amount = round(np.random.uniform(850, 85000), 2)
    dt = pd.Timestamp(year=2025, month=random.randint(1,6), day=random.randint(1,28))
    rows.append({
        "Contracting Authority": random.choice(authorities),
        "Supplier Name": supplier,
        "Amount EUR": random.choice([amount, f"{amount:.2f}", f"{amount:,.2f}".replace(",", " ").replace(".", ",")]),
        "Date": random.choice([dt.strftime("%Y-%m-%d"), dt.strftime("%d.%m.%Y"), dt.strftime("%d/%m/%Y")]),
        "Contract Type": random.choice(contract_types),
        "Procedure No": f"IEP-2025-{random.randint(100,999)}",
        "Supplier Canonical (for teacher)": canon
    })
save_df_xlsx(pd.DataFrame(rows), base / "2_procurement_data" / "procurement_raw.xlsx", sheet_name="Iepirkumi")

# 3. Employee registry
positions = {
    "IT nodaļa": ["Sistēmu administrators", "Datu analītiķis", "Projektu vadītājs"],
    "Personāla nodaļa": ["Personāla speciālists", "Vecākais speciālists"],
    "Finanšu nodaļa": ["Grāmatvedis", "Finanšu analītiķis", "Budžeta plānotājs"],
    "Iepirkumu nodaļa": ["Iepirkumu speciālists", "Vecākais iepirkumu speciālists"],
    "Juridiskā nodaļa": ["Jurists", "Vecākais jurists"],
    "Klientu apkalpošana": ["Klientu konsultants", "Pakalpojumu koordinators"],
    "Attīstības nodaļa": ["Attīstības projektu koordinators", "Datu speciālists"],
    "Izglītības pārvalde": ["Izglītības speciālists", "Programmu koordinators"]
}
rows = []
for i in range(95):
    dep = random.choice(departments_lv)
    salary = int(np.random.normal(1850, 420))
    salary = max(1100, min(salary, 3600))
    start = pd.Timestamp(year=random.randint(2018, 2025), month=random.randint(1,12), day=random.randint(1,28))
    rows.append({
        "Employee ID": f"EMP-{1000+i}",
        "Name": synth_name(),
        "Department": dep,
        "Position": random.choice(positions[dep]),
        "Salary": salary,
        "FTE": random.choice([1.0, 1.0, 1.0, 0.8, 0.75]),
        "Start Date": start.strftime("%Y-%m-%d"),
        "Municipality": random.choice(municipalities[:12])
    })
save_df_xlsx(pd.DataFrame(rows), base / "3_employee_registry" / "employees.xlsx", sheet_name="Darbinieki")

# 4. Municipality data
population_rows = []
budget_rows = []
for m in municipalities:
    pop = int(np.random.uniform(7000, 640000)) if m != "Rīga" else 605000
    bud = int(pop * np.random.uniform(900, 2400))
    population_rows.append({"Municipality": m, "Population": pop, "Reference Year": 2025})
    budget_rows.append({
        "Municipality": m,
        "Budget": bud,
        "Capital Expenditure": int(bud * np.random.uniform(0.12, 0.28)),
        "Reference Year": 2025
    })
save_df_xlsx(pd.DataFrame(population_rows), base / "4_municipality_data" / "population.xlsx", sheet_name="Population")
save_df_xlsx(pd.DataFrame(budget_rows), base / "4_municipality_data" / "budget.xlsx", sheet_name="Budget")

# 5. Mixed quality data
wb = Workbook()
ws = wb.active
ws.title = "RawData"
messy_rows = []
for _ in range(70):
    messy_rows.append([
        random.choice(municipalities[:10]),
        random.choice(["IT nodaļa", "Dept: IT nodaļa", "Finanšu nodaļa", "DEPARTMENT: Finanšu nodaļa", "Personāla nodaļa"]),
        random.choice(["Apmācības", "Programmatūra", "Biroja preces", "Konsultācijas"]),
        random.choice([
            round(np.random.uniform(50, 9999), 2),
            str(round(np.random.uniform(50, 9999), 2)),
            f"{round(np.random.uniform(50, 9999), 2):.2f}".replace(".", ",")
        ]),
        random.choice(["2025-01-15", "15.02.2025", "03/03/2025", "", None]),
        random.choice(["Jā", "Nē", "ja", "ne", None])
    ])
ws["A1"] = "Eksporta pārskats"
ws["A2"] = "Izveidots: 2025-06-30"
ws.append([])
ws.append(["Municipality", "Dept", "Category", "Amount", "Date", "Approved"])
for row in messy_rows[:25]:
    ws.append(row)
ws.append([])
ws.append(["Municipality", "DEPARTMENT", "Category", "Amount", "Date", "Approved"])
for row in messy_rows[25:]:
    ws.append(row)
for extra in messy_rows[:5]:
    ws.append(extra)
for col in "ABCDEF":
    ws.column_dimensions[col].width = 22
wb.save(base / "5_mixed_quality_data" / "mixed_data.xlsx")

# 6. Reporting template
wb = Workbook()
ws = wb.active
ws.title = "Kopsavilkums"
ws["A1"] = "Automatizētais pārskats"
ws["A1"].font = Font(bold=True, size=14)
ws["A3"] = "Iestāde:"
ws["A4"] = "Periods:"
ws["A6"] = "Kategorija"
ws["B6"] = "Summa EUR"
for cell in ws[6]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 14
ws2 = wb.create_sheet("Dati")
ws2["A1"] = "Šeit ielīmēt vai eksportēt apstrādātos datus"
ws2["A1"].font = Font(bold=True)
ws2.column_dimensions["A"].width = 40
wb.save(base / "6_reporting_template" / "report_template.xlsx")

# 7. Chart data
summary_df = pd.DataFrame(budget_summary_rows).groupby(["Month","Department"], as_index=False)["Amount"].sum()
pivot = summary_df.pivot(index="Month", columns="Department", values="Amount").fillna(0)
pivot = pivot.reindex([m[1].capitalize() for m in months])
chart_file = base / "7_chart_data" / "monthly_summary.xlsx"
save_df_xlsx(pivot.reset_index().rename(columns={"index":"Month"}), chart_file, sheet_name="Summary")

wb = load_workbook(chart_file)
ws = wb["Summary"]
chart = BarChart()
chart.title = "Mēneša izdevumi pa nodaļām"
chart.y_axis.title = "EUR"
chart.x_axis.title = "Mēnesis"
data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 16
ws.add_chart(chart, "J2")
wb.save(chart_file)

# README
(base / "README.txt").write_text(
"""Latvia Government Context Excel Practice Datasets

Synthetic but realistic Excel files for Python-based Excel automation training.

Folders:
1_budget_reports
2_procurement_data
3_employee_registry
4_municipality_data
5_mixed_quality_data
6_reporting_template
7_chart_data

All data is synthetic. No real personal data is included.
""",
    encoding="utf-8"
)

zip_path = Path("latvia_gov_excel_datasets.zip")
if zip_path.exists():
    zip_path.unlink()

with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
    for p in base.rglob("*"):
        zf.write(p, p.as_posix())

print(f"Created: {zip_path}")